from openpyxl import load_workbook
from chinese_stroke_sorting import sort_by_stroke
# import sys
# from openpyxl.styles import Color, PatternFill, Font, Border
# import openpyxl

charNoDict = {}

def readTextFileLines(path):
    file = open(path,"r")
    return file.readlines()

def genCharTable():
    lines = readTextFileLines('crystalIMEtable.asm')
    key = ''
    index = 0
    for line in lines:
        if ':' in line:
            key = line.replace(':','').replace('\n','')
            charNoDict[key] = {}
            index = 0
        else:
            modLine = line.replace('\n','')
            for char in modLine:
                # print(char)
                # print(char.encode('utf-8'))
                charNoDict[key][char] = index
                index += 1


def addAtSymbol(input):
    output = input
    length = len(input)
    for i in range(7-length):
        output += '@'
    return output

def modHexText(input):
    output = input
    output = output.replace('HEX','$')
    outputlist = list(output)
    outputlist.insert(3, ',')
    outputlist.insert(4, '$')
    output = ''.join(outputlist)
    output += ','
    return output

def sortChar(key,array):
    input = array
    tmp = []
    tmp2 = []
    # print(charNoDict[key])
    # print(input)
    for item in array:
        # print('item')
        # print(item)
        # print(item.encode('utf-8'))
        if key in charNoDict:
            if item in charNoDict[key]:
                # print('Found: ' + item)
                tmp.append(item)
            else:
                # print('Not: ' + item)
                tmp2.append(item)
        else:
            tmp = array

    if key in charNoDict:
        tmp = sorted(tmp,key=lambda x:(charNoDict[key][x]))
    # print(tmp)
    # print(input)
    output = tmp + tmp2
    return output

wb = load_workbook(filename = 'char.xlsx')
genCharTable()
# print(charNoDict)
# charNo = readTextFileLines('bh.txt')
# charNoDict = {}
# for line in charNo:
#     component = line.split('\t')
#     if len(component) >= 2:
#         if not component[0] in charNoDict:
#             charNoDict[component[0]] = int(component[1])

codeTable = {}
for sheet in wb._sheets:
    row = 1
    while sheet.cell(row=row, column=6).value != 'end':
        char = sheet.cell(row=row, column=6).value
        hexStr = modHexText(sheet.cell(row=row, column=7).value)
        codeTable[char] = hexStr
        row += 1

charTable = {}
tmpNum = 0
for sheet in wb._sheets:
    row = 1
    while sheet.cell(row=row, column=1).value != 'end':
        for col in range(2,6):
            pin = sheet.cell(row=row, column=col).value
            char = sheet.cell(row=row, column=1).value
            if pin != None:
                if not pin in charTable:
                    charTable[pin] = [char]
                    tmpNum += 1
                else:
                    charTable[pin].append(char)
                    tmpNum += 1
        row += 1
# print('tmpnum')
# print(tmpNum)
total = 0
for key in charTable:
    # if key.upper() == '***':
    #     print(key)
    print('IME_'+key.upper()+'_Char:')
    text = '\t db '
    chs = '\t ; '
    # charTable[key] = sort_by_stroke(charTable[key])
    charTable[key] = sortChar(key.upper(),charTable[key])
    # charTable[key] = sorted(charTable[key])
    # charTable[key] = sorted(charTable[key],key=lambda x:(charNoDict[x]))
    for i in range(len(charTable[key])):
        item = charTable[key][i]
        code = codeTable[item]
        text += code
        chs += item + ' '
    text += '$50'
    
    print(text)
    print(chs)
    print('\t ; ' + str(len(charTable[key])))
    total += len(charTable[key])

print('\t db $50,$50,$50,$50,$50,$50,$50,$50,$50,$50,$50,$50,$50,$50,$50,$50,$50,$50,$50,$50,$50,$50,$50,$50,$50')
print('\t ; ' + str(total))