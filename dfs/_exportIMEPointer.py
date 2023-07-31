from openpyxl import load_workbook
# import sys
# from openpyxl.styles import Color, PatternFill, Font, Border
# import openpyxl

def addAtSymbol(input):
    output = input
    length = len(input)
    for i in range(7-length):
        output += '@'
    return output

wb = load_workbook(filename = 'char.xlsx')

dicts =[{},{},{},{},{},{}] 

for sheet in wb._sheets:
    row = 1
    while sheet.cell(row=row, column=1).value != 'end':
        for col in range(2,6):
            pin = sheet.cell(row=row, column=col).value
            if pin != None:
                length = len(pin)
                if not pin in dicts[length - 1]:
                    dicts[length - 1][pin] = 1
                else:
                    dicts[length - 1][pin] += 1
        row += 1

print('PinyinStrTableInfoTable:')
for i in range(0,6):
    dict = dicts[i]
    print('\tdb ' + str(len(dict)))

for i in range(0,6):
    dict = dicts[i]
    print('PinyinStrTableLen'+ str(i + 1) + ':')
    # for key in dict:
    #     print('\tdb \"'+ addAtSymbol(key).upper() + '\"')
    #     print('\tdw IME_'+ key.upper() + '_Char')
    #     count = dict[key]
    #     pagenum = int(count / 12)
    #     if count % 12 != 0:
    #         pagenum += 1
    #     print('\tdb '+ str(pagenum))
    #     print('\t;'+ str(count))
        # if count >= 90:
        #     print('TOO MANY: ' + str(count))
    print()


newList = []
for i in range(0,6):
    dict = dicts[i]
    for key in dict:
        newList.append((key,dict[key]))
newList = sorted(newList,key=lambda x:(x[0]))

total = 0
itemCount = 0
for item in newList:
    itemCount += 1
    print('\tdb \"'+ addAtSymbol(item[0]).upper() + '\"')
    print('\tdw IME_'+ item[0].upper() + '_Char')
    count = item[1]
    pagenum = int(count / 12)
    if count % 12 != 0:
        pagenum += 1
    print('\tdb '+ str(pagenum))
    print('\t;'+ str(count))
    total += count
    # if count >= 90:
    #     print('TOO MANY: ' + str(count))
print('\t')
print('\tdb \"@@@@@@@\"')
print('\tdw IMEBlank')
print('\tdb 1')
print('; total Pinyin: ' + str(itemCount))
print('; total: ' + str(total))