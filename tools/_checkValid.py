from openpyxl import load_workbook
import sys
from openpyxl.styles import Color, PatternFill, Font, Border
import openpyxl
from enum import Enum
import os

warningFill = PatternFill(start_color='0000FFFF',
                   end_color='0000FFFF',
                   fill_type='solid')
noFill = openpyxl.styles.PatternFill(fill_type=None)
textReplacement = {'#MON':'#','#':'宝可梦','&':'训练家'}

def replaceText(text,dict):
    output = removeNone(text)
    for key in dict:
        output = output.replace(key,dict[key])
    return output

class bcolors:
    HEADER = '\033[95m'
    OKBLUE = '\033[94m'
    OKCYAN = '\033[96m'
    OKGREEN = '\033[92m'
    WARNING = '\033[93m'
    FAIL = '\033[91m'
    ENDC = '\033[0m'
    BOLD = '\033[1m'
    UNDERLINE = '\033[4m'
print(bcolors.OKGREEN)

def removeNone(text):
    if text == None:
        return ''
    return text

def isChinese(_char):
    if _char == '…' or _char == '　' or _char == '！' or _char == '？' or _char == '：' or _char == '。' or _char == '，':
        return True
    return '\u4e00' <= _char <= '\u9fa5'

def ifContainsChinese(text):
    input = removeNone(text)
    for theChar in input:
        if isChinese(theChar):
            return True
    return False

def ifOverLength(text,maxPixels):
    if text == None:
        return False
    newText = text.replace('<PLAYER>','PLAYER ').replace('<RIVAL>','RIVALN ').replace('\n','')
    if len(newText) <= 0:
        return False
    charProp = isChinese(newText[0])
    length = 1
    pixels = 0
    for i in range(1,len(newText)):
        character = newText[i]
        if character == '@':
            continue
        newProp = isChinese(character)
        if newProp == charProp:
            length += 1
        else:
            if charProp:
                chsParts = float(int(length / 2))
                if length % 2 != 0:
                    chsParts += 16.0/24.0
                pixels += chsParts * 24
            else:
                pixels += length * 8
            charProp = newProp
            length = 0
        
        if i == len(newText) - 1:
            if charProp:
                chsParts = float(int(length / 2))
                if length % 2 != 0:
                    chsParts += 16.0/24.0
                pixels += chsParts * 24
            else:
                pixels += length * 8
    # print(pixels)
    return pixels > maxPixels

halfNumChars = ['0','1','2','3','4','5','6','7','8','9']
def ifTextContains(text,list):
    tmp = removeNone(text)
    for item in list:
        if item in tmp:
            return True
    return False

class InfoType(Enum):
    ERROR = 1
    WARNING = 2
    INFO = 3

class Instruction():
    label = ''
    inst = ''
    content = ''
    filePath = ''
    sheetTitle = ''

def printLog(type,sheet,instuction,message):
    global errors
    global warnings
    global infos
    if type == InfoType.ERROR:
        print(bcolors.FAIL)
        errors += 1
    elif type == InfoType.WARNING:
        print(bcolors.WARNING)
        warnings += 1
    else:
        print(bcolors.OKCYAN)
        infos += 1
    print('xlsx 路径：' + xlsxListPath)
    print('sheet 标题：' + sheet.title)
    if instuction != None:
        print(instuction.label)
        print(instuction.inst + ' ' + instuction.content)
    print(message)
    print(bcolors.OKGREEN)

errors = 0
warnings = 0
infos = 0
textOldPlacerCommands=['TX_RAM','TX_NUM','TX_BCD']
textPlacerCommands=['text_ram','text_decimal','text_bcd']
def getInstDict(col,sheet,filePath):
    outputDict = {}
    id = 2
    labelrows = []
    labels = []
    while sheet.cell(row=id, column=col).value != 'end' and id <= 10000:
        label = removeNone(sheet.cell(row=id, column=col).value)
        if ':' in label:
            labelrows.append(id)
            labels.append(label)
            if sheet.cell(row=id, column=col + 1).value != None:
                 printLog(InfoType.WARNING,sheet,None,label + '\ncol ' + str(col) + '\n旁边有多余内容！')
        id += 1
    labelrows.append(id) 
    if sheet.cell(row=id, column=col + 1).value != None:
        printLog(InfoType.WARNING,sheet,None,'row '+ str(id) + '\ncol ' + str(col) + '\n旁边有多余内容！')
    if id >= 10000:
        printLog(InfoType.ERROR,sheet,None,sheet.title + '\ncol ' + str(col) + '\n找不到 end 结束符号！')

    for i in range(len(labelrows) - 1):
        instructions = []
        for j in range(labelrows[i],labelrows[i + 1]):
            inst = sheet.cell(row=j, column=col + 1).value
            content = sheet.cell(row=j, column=col + 2).value
            if inst != None and not ';' in inst:
                newInstructions = Instruction()
                newInstructions.label = labels[i]
                newInstructions.inst = inst
                newInstructions.content = removeNone(content)
                newInstructions.sheetTitle = sheet.title
                newInstructions.filePath = filePath
                instructions.append(newInstructions)
        #         if ifTextContains(inst,textOldPlacerCommands):
        #             printLog(InfoType.ERROR,sheet,newInstructions,'有非法内容！')
        # if len(instructions) == 0:
        #     printLog(InfoType.INFO,sheet,None,labels[i] + '\n col' + str(col) + '\n标签为空白指令')
        outputDict[labels[i]] = instructions
    
    # for key in outputDict:
    #     print('key: ' + key)
    #     instuctions = outputDict[key]
    #     for instuction in instuctions:
    #         print(instuction.label)
    #         print(instuction.inst + ' ' + instuction.content)
    #     print()
    return outputDict

textStarterCommands=['text','text_start','text_ram','text_decimal','text_bcd','text $4c,','text "<_CONT>@"','vc_patch Change_link_closed_inactivity_message ']
textEndingCommands=['done','prompt','dex','text_end']

def ifTextIsInList(text,commands):
    for command in commands:
        if text == command:
            return True
    return False

def checkDictValid(instDict):
    for key in instDict:
        instructions = instDict[key]

        
        if len(instructions) > 0:
            # 检查开头和结尾符号
            if not ifTextIsInList(instructions[0].inst,textStarterCommands):
                printLog(InfoType.ERROR,sheet,instructions[0],'开头不合法！')
            if not ifTextIsInList(instructions[-1].inst,textEndingCommands):
                if not '@@' in instructions[-1].content:
                    printLog(InfoType.ERROR,sheet,instructions[-1],'结尾不合法！')

            # 检查战斗后文本长度
            if 'EndBattleText' in instructions[0].label:
                lengthchk = replaceText(instructions[0].content,textReplacement)
                if ifOverLength(lengthchk,8*8):
                    printLog(InfoType.WARNING,sheet,instructions[0],'战斗后文本可能太长！')
        
        for i in range(0,len(instructions)):
            instruction = instructions[i]

            # 检查一般文本长度
            if not ifTextIsInList(instruction.inst,textPlacerCommands):
                lengthchk = replaceText(instruction.content,textReplacement)
                if ifOverLength(lengthchk,18 * 8):
                    printLog(InfoType.WARNING,sheet,instruction,'文本可能太长！')

            # 检查 textPlacerCommands=['text_ram','text_decimal','text_bcd'] 是否合法
            if i > 0:
                if ifTextIsInList(instruction.inst,textPlacerCommands):
                    if len(instructions[i - 1].content) > 0:
                        if instructions[i - 1].content[-1] != '@':
                            printLog(InfoType.ERROR,sheet,instruction,'上一条结尾不是 @ 符号！')
                    else:
                        printLog(InfoType.ERROR,sheet,instruction,'上一条结尾不是 @ 符号！')

            # 检查是否是老文本
            if '@@' in instruction.content:
                printLog(InfoType.INFO,sheet,instruction,'发现旧工程文本@@！')
            
            # if instruction.inst == 'text' and instruction.content == '':
            #     printLog(InfoType.INFO,sheet,instruction,'发现旧工程text空白符号！')
            
            #其他检查
            if ifTextContains(instruction.content,halfNumChars) and not ifTextIsInList(instruction.inst,textPlacerCommands):
                printLog(InfoType.INFO,sheet,instruction,'发现半角数字！')


def compareDicts(origDict,transDict):
    for key in origDict:
        if not key in transDict:
            printLog(InfoType.ERROR,sheet,None,key + '：标签不存在！')
        else:
            origInstructions = origDict[key]
            transInstructions = transDict[key]
            if len(origInstructions) > 0 and len(transInstructions) > 0:
                inst1 = origInstructions[-1].inst.replace('text_end','text')
                inst2 = origInstructions[-1].inst.replace('text_end','text')
                if inst1 != inst2:
                     printLog(InfoType.ERROR,sheet,origInstructions[-1],'结尾指令不一致！')
            else:
                printLog(InfoType.INFO,sheet,None,key +'：标签为空白指令')

xlsxListPath = sys.argv[1]
colOrig = int(sys.argv[2])
colTrans = int(sys.argv[3])
bypassFileCheck = int(sys.argv[4])


print(xlsxListPath +'：检查 xlsx 合法性...')
print()

wb = load_workbook(filename = xlsxListPath)
for sheet in wb._sheets:
    outputPath = sheet.cell(row=1, column=1).value
    if not os.path.isfile(outputPath):
        if bypassFileCheck == 0:
            printLog(InfoType.ERROR,sheet,None,outputPath + '\n文件不存在！')
    origDict = getInstDict(colOrig,sheet,xlsxListPath)
    TransDict = getInstDict(colTrans,sheet,xlsxListPath)
    checkDictValid(TransDict)
    compareDicts(origDict,TransDict)

print('检查结束')
print(bcolors.FAIL)
print('发现错误：'+ str(errors))
print(bcolors.WARNING)
print('发现警告：'+ str(warnings))
print(bcolors.OKCYAN)
print('发现提醒：'+ str(infos))