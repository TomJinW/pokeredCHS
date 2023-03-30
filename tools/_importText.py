from openpyxl import load_workbook
import sys
from openpyxl.styles import Color, PatternFill, Font, Border
import openpyxl
warningFill = PatternFill(start_color='0000FFFF',
                   end_color='0000FFFF',
                   fill_type='solid')
noFill = openpyxl.styles.PatternFill(fill_type=None)
import charmap

class bcolors:
    HEADER = '\033[95m'
    OKBLUE = '\033[94m'
    OKCYAN = '\033[96m'
    OKGREEN = '\033[92m'
    WARNING = '\033[93m'
    FAIL = '\033[91m'
    ENDC = '\033[0m'
    BOLD = '\033[1m'
    UNDERLINE = '\033[4m'

xlsxListPath = sys.argv[1]
mode = int(sys.argv[2])
extraTextPath = ''
if len(sys.argv) > 3:
    extraTextPath = sys.argv[3]

wb = load_workbook(filename = xlsxListPath)

def readLines(filename):
    fileList = []
    with open(filename) as f:
        fileList = f.readlines()
        return fileList
    
def removeNone(text):
    if text == None:
        return ''
    return text

def isChinese(_char):
    if _char == '…' or _char == '　' or _char == '！' or _char == '？' or _char == '：' or _char == '。':
        return True
    return '\u4e00' <= _char <= '\u9fa5'

def ifContainsChinese(text):
    input = removeNone(text)
    for theChar in input:
        if isChinese(theChar):
            return True
    return False

def ifOverLength(text,maxPixels):
    if text == None:
        return False
    newText = text.replace('<PLAYER>','PLAYER ').replace('<RIVAL>','RIVALN ').replace('\n','')
    if len(newText) <= 0:
        return False
    charProp = isChinese(newText[0])
    length = 1
    pixels = 0
    for i in range(1,len(newText)):
        character = newText[i]
        if character == '@':
            continue
        newProp = isChinese(character)
        if newProp == charProp:
            length += 1
        else:
            if charProp:
                chsParts = float(int(length / 2))
                if length % 2 != 0:
                    chsParts += 16.0/24.0
                pixels += chsParts * 24
            else:
                pixels += length * 8
            charProp = newProp
            length = 0
        
        if i == len(newText) - 1:
            if charProp:
                chsParts = float(int(length / 2))
                if length % 2 != 0:
                    chsParts += 16.0/24.0
                pixels += chsParts * 24
            else:
                pixels += length * 8
    # print(pixels)
    return pixels > maxPixels
    
def getLabelType(label):
    if label == None:
        return -1
    elif label[-1] == ':':
        return 0
    elif label[0] == ';':
        components = label.split(' ')
        if components[-1].lower() == 'start':
            return 3

# ifOverLength('我大意了。',64)
# exit(0)
textCommands=['page','text','line','cont','para','next']
textPlacerCommands=['text_ram','text_decimal','text_bcd']
# charmap "%",$2EB8 ;PP
# charmap "&", $5D ;训练家
# charmap "+", $5E ;火箭队
# charmap "^", $5C ;招式学习器
# charmap "~", $56 ;......
textReplacement = {'#MON':'#','#':'宝可梦','&':'训练家'}
#textRevReplacement = {'宝可梦':'#','训练家':'&','火箭队':'+','#MON':'#','招式学习器':'^','……':'~'}
textRevReplacement = {'<PLAYER>':'ć','<RIVAL>':'č','宝可梦':'#','&':'训练家','#MON':'#','+':'火箭队'}
def getInstType(inst):
    if inst == None:
        return -1
    for cmd in textCommands:
        if cmd in inst:
            return 0
    for cmd in textPlacerCommands:
        if inst == cmd:
            return 1
    return -2

def replaceText(text,dict):
    output = removeNone(text)
    for key in dict:
        output = output.replace(key,dict[key])
    return output


def textFormat(text,mode):
    output = replaceText(removeNone(text),textRevReplacement)
    if mode == 0:
        return '\"' + output + '\"'
    return output

charMap = charmap.readCharMaps()
extraList = []
if extraTextPath != '':
    
    lines = readLines(extraTextPath)
    textLine = ''
    labelStarted = False
    for line in lines:
        components = line.replace('\n','').split(' ')

        
        if labelStarted:
            textLine += line
            if line[0] == ';' and len(components) > 0:
                if components[-1].lower() == 'end':
                    labelStarted = False
                    extraList.append(textLine)
                    textLine = ''
        if line[0] == ';' and len(components) > 0:
            if components[-1].lower() == 'start':
                labelStarted = True
                textLine += line

halfChars = ['0','1','2','3','4','5','6','7','8','9']
def ifcontains(text):
    if text == None:
        return False
    for sChar in halfChars:
        if sChar in text:
            return True

waringText = ''
for sheet in wb._sheets:
    # init setup
    outputText = ""
    outputPath = sheet.cell(row=1, column=1).value
    
    id = 2
    extraID = 0
    while sheet.cell(row=id, column=mode).value != 'end' and id <= 10000:
        label = sheet.cell(row=id, column=mode).value
        labelType = getLabelType(label)

        if labelType == 0:
            outputText += label + '\n'
            if 'EndBattleText'.upper() in label.upper():
                text = sheet.cell(row=id+1, column=mode + 2).value
                if ifOverLength(text,8*8):
                    sheet.cell(row=id+1, column=mode + 2).fill = warningFill
                    print(bcolors.WARNING + '警告！战斗结束文本 in')
                    print(xlsxListPath)
                    print(label)
                    print(text)
                    print('可能过长！\n')
                # else:
                #     sheet.cell(row=id+1, column=mode + 2).fill = noFill

        elif labelType == 3:
            # get extra content
            outputText += extraList[extraID]
            extraID += 1
            a = 1
        else:
            inst = sheet.cell(row=id, column=mode + 1).value
            content = sheet.cell(row=id, column=mode + 2).value

            instType = getInstType(inst)
            if instType == 0 or instType == 1:
                textline = textFormat(content,instType)
                if ifContainsChinese(textline):
                    textline = charmap.replaceText(textFormat(content,1),charMap)
            
                outputText += '\t' + inst + ' ' + textline.replace('\"\"','') + '\n'
                lengthchk = replaceText(content,textReplacement)
                if instType == 0:
                    if ifOverLength(lengthchk,18 * 8):
                        # waringText += 'Warning\n'
                        # waringText += xlsxListPath + '\n'
                        # waringText += sheet.title + '\n'
                        # waringText += replaceText(content,textReplacement) + '\n文本可能过长！\n' + '\n'
                        print(bcolors.FAIL + '警告！')
                        print(xlsxListPath)
                        print(sheet.title)
                        print('' + replaceText(content,textReplacement) + '\n文本可能过长！\n')
                        sheet.cell(row=id, column=mode + 2).fill = warningFill
                    # else:
                    #     sheet.cell(row=id, column=mode + 2).fill = noFill

                    if ifcontains(content):
                        print(bcolors.OKBLUE + '提醒！')
                        print(xlsxListPath)
                        print(sheet.title)
                        print('' + replaceText(content,textReplacement) + '\n含有半角符号！\n')
                if instType == 1:
                    lastContent = sheet.cell(row=id - 1, column=mode + 2).value
                    if lastContent == None:
                        lastContent = ''
                    lastInst = sheet.cell(row=id - 1, column=mode + 1).value
                    if (lastInst != None and not '@' in lastContent):
                        print(bcolors.FAIL + '警告！')
                        print(id - 1)
                        print(xlsxListPath)
                        print(inst)
                        print(content)
                        print(lastContent)
                        print('未找到 @ 符号！')
                        print('')
                        sheet.cell(row=id - 1, column=mode + 2).fill = warningFill

            elif instType == -2:
                outputText += '\t' + inst + '\n'
            else:
                outputText += '\n'
        id += 1
    
    with open(outputPath, 'w') as f:
        f.write(outputText)

wb.save(xlsxListPath)
   



        





