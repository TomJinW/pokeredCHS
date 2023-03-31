from openpyxl import load_workbook
from openpyxl.styles import Color, PatternFill, Font, Border

import sys
import openpyxl
import shutil
import os
import charmap

class bcolors:
    HEADER = '\033[95m'
    OKBLUE = '\033[94m'
    OKCYAN = '\033[96m'
    OKGREEN = '\033[92m'
    WARNING = '\033[93m'
    FAIL = '\033[91m'
    ENDC = '\033[0m'
    BOLD = '\033[1m'
    UNDERLINE = '\033[4m'

def addFilePaths(path):
    if not path in filePaths:
        filePaths.append(path)
        os.makedirs(os.path.dirname(tmp+path), exist_ok=True)
        

def readTextFile(path):
    file = open(path,"r")
    return file.read()

def removeNone(text):
    if text == None:
        return ''
    return text

tmp = 'tmp/'
# Program Start
xlsxListPath = sys.argv[1]
mode = int(sys.argv[2])

# Load Workbook
wb = load_workbook(filename = xlsxListPath)

filePaths = []

charMap = charmap.readCharMaps()

print()
print(bcolors.OKGREEN)
print('Importing db Data...')

for sheet in wb._sheets:
    replacees = []
	# Get init File Path
    filePath = sheet.cell(row=1, column=mode).value
    endChar = removeNone(sheet.cell(row=1, column=mode+1).value)
	# Backup Original Files
    addFilePaths(filePath)
    text2Modify = readTextFile(filePath)

    id = 2
    while sheet.cell(row=id, column=mode).value != 'end':

        if sheet.cell(row=id, column=mode).value != None:
            if sheet.cell(row=id, column=mode).value != 'end':
                with open(filePath, 'w') as f:
                    f.write(text2Modify)
                replacees = []
                filePath = sheet.cell(row=id, column=mode).value
                endChar = removeNone(sheet.cell(row=id, column=mode+1).value)
                addFilePaths(filePath)
                text2Modify = readTextFile(filePath)
                id += 1
                continue

        replacee = removeNone(sheet.cell(row=id, column = mode + 1).value)
        replacer = removeNone(sheet.cell(row=id, column = mode + 2).value) + endChar

        if replacee != '':
            repetitive = False
            for currentsReplacee in replacees:
                if replacee in currentsReplacee:
                    repetitive = True
                    print(bcolors.OKBLUE + "Warning! Duplicate: " + replacee)
            if not repetitive:
                replacees.append(replacee)    

            if sheet.cell(row=id, column = mode + 3).value != None:
                newReplacee = sheet.cell(row=id, column = mode + 3).value
                text2Modify = text2Modify.replace(replacee,newReplacee)
            else:
                text2Modify = text2Modify.replace(replacee,charmap.replaceText(replacer,charMap))


        id += 1

    # print(text2Modify)
    with open(filePath, 'w') as f:
        f.write(text2Modify)
# input(bcolors.OKGREEN + "Any key..")
print(bcolors.OKGREEN)
print()
print('db Data Import complete.')
input("Press Return to proceed..")



